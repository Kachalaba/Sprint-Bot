"""Tests for export service producing PB/SoB data."""

from __future__ import annotations

import asyncio
import csv
import sqlite3
from datetime import datetime, timezone
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook

from services.export_service import ExportService


def _setup(tmp_path: Path) -> Path:
    db_path = tmp_path / "export.db"
    with sqlite3.connect(db_path) as conn:
        conn.executescript(
            """
            CREATE TABLE results (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                athlete_id INTEGER NOT NULL,
                athlete_name TEXT NOT NULL DEFAULT '',
                stroke TEXT NOT NULL,
                distance INTEGER NOT NULL,
                total_seconds REAL NOT NULL,
                timestamp TEXT NOT NULL,
                is_pr INTEGER NOT NULL DEFAULT 0
            );
            CREATE TABLE result_segments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                result_id INTEGER NOT NULL,
                segment_index INTEGER NOT NULL,
                split_seconds REAL NOT NULL
            );
            """
        )
        now = datetime(2024, 6, 1, tzinfo=timezone.utc)
        cursor = conn.execute(
            """
            INSERT INTO results (athlete_id, athlete_name, stroke, distance, total_seconds, timestamp, is_pr)
            VALUES (?,?,?,?,?,?,0)
            """,
            (5, "Export", "freestyle", 100, 62.0, now.isoformat()),
        )
        result_id = cursor.lastrowid
        for idx, value in enumerate((15.5, 15.4, 15.5, 15.6)):
            conn.execute(
                """
                INSERT INTO result_segments (result_id, segment_index, split_seconds)
                VALUES (?, ?, ?)
                """,
                (result_id, idx, value),
            )
        conn.commit()
    return db_path


def test_export_service_csv_and_excel(tmp_path: Path) -> None:
    db_path = _setup(tmp_path)
    service = ExportService(db_path)
    csv_bytes = asyncio.run(service.export_pb_sob([5], fmt="csv"))
    text = csv_bytes.decode("utf-8")
    reader = csv.reader(StringIO(text))
    rows = list(reader)
    assert len(rows) > 1
    assert rows[0][0] == "athlete_id"

    xlsx_bytes = asyncio.run(service.export_pb_sob([5], fmt="xlsx"))
    workbook = load_workbook(BytesIO(xlsx_bytes))
    sheet = workbook.active
    assert sheet.max_row >= 2
